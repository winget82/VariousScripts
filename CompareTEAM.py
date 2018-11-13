# -*- coding: utf-8 -*-

import pyperclip as clip
import pyautogui as p
import os
import time
import openpyxl
import re
import csv
from itertools import zip_longest

dgntext = []
dgntext2 = []
dgntext3 = []
xcltext = []
xclfound = []
dgnfound = []
unfounddgntxt = []
unfounddgntxt2 = []
unfoundxcltxt = []

#change directory to where images for pyautogui are saved
os.chdir('C:\\Users\\...')

#set up to search for 4 digits together
finddigits = re.compile('\d\d\d\d')
#function to execute the 4 digit search
def contains_digits(d):
    return bool(finddigits.search(d))

location = input('Are you working from the (H)ome or (O)ffice?')

#FROM OFFICE LOCATION (DGN PORTION)
if location.upper() == 'O':

    p.click(875,9)

    try:
        #maximize microstation view window
        maxwin = p.locateOnScreen('maxwin.PNG', grayscale=True)
        maxwin_center = p.center(maxwin)
        p.moveTo(maxwin_center, duration=1)
        p.click()
    except:
        print("maximize window button not found")

    #zoom to extents
    zoomext = p.locateOnScreen('zoomext.PNG', grayscale=True)
    zoomext_center = p.center(zoomext)
    p.moveTo(zoomext_center, duration=1)
    print("zooming to extents")
    p.click()

    time.sleep(1)

    #select tool
    select = p.locateOnScreen('select.PNG', grayscale=True)
    select_center = p.center(select)
    p.moveTo(select_center, duration=1)
    print("selecting elements")
    p.click()

    #drag box to select everything in the dgn
    p.moveTo(42, 293)
    p.click()
    time.sleep(1)
    p.dragTo(1495, 979, 2, button='left')

    #copy to clipboard the text (will have excess stuff not needed)
    p.hotkey('ctrl', 'c') #this is for a physical copy to clip board

#FROM HOME LOCATION (DGN PORTION)
elif location.upper() == 'H':

    p.click(1324,7)

    try:
        #maximize microstation view window
        maxwinH = p.locateOnScreen('maxwinH.PNG', grayscale=True)
        maxwinH_center = p.center(maxwinH)
        p.moveTo(maxwinH_center, duration=1)
        p.click()
    except:
        print("maximize window button not found")

    #zoom to extents
    zoomextH = p.locateOnScreen('zoomextH.PNG', grayscale=True)
    zoomextH_center = p.center(zoomextH)
    p.moveTo(zoomextH_center, duration=1)
    print("zooming to extents")
    p.click()

    time.sleep(1)

    #select tool
    selectH = p.locateOnScreen('selectH.PNG', grayscale=True)
    selectH_center = p.center(selectH)
    p.moveTo(selectH_center, duration=1)
    print("selecting elements")
    p.click()

    #drag box to select everything in the dgn
    p.moveTo(43, 220)
    p.click()
    time.sleep(1)
    p.dragTo(1470, 991, 2, button='left')

    #copy to clipboard the text (will have excess stuff not needed)
    p.hotkey('ctrl', 'c') #this is for a physical copy to clip board
    
#CONTINUED DGN PORTION (MUTUAL HOME OR OFFICE)    
time.sleep(1)

text = clip.paste() #this will paste the copied text into a variable

#print(text)
dgntext = text.split('\n')
#print(dgntext)

for i in dgntext:
    i = i.replace('\r', '')
    dgntext2.append(i)
#print(dgntext2)

for i in dgntext2:
    if contains_digits(i) == True:
        dgntext3.append(i)
#print(dgntext3)

#BELOW THIS IS HOW TO GET CELL VALUES OF A WORKSHEET WITHIN AN XLSX WORKBOOK (iterates through all the values in
# columns B and C and joins them)

from openpyxl import load_workbook
nameFile = input('What would you like to name your file? ')
file = input('Excel file path and file name?')
Sheet_name = input('Excel sheet name?').upper()
workbook = load_workbook(file, data_only=True) #CHANGE TO YOUR FILENAME
sheet = workbook[Sheet_name] #CHANGE TO THE SHEETNAME OF THE PAGE YOU WANT - MAY WANT TO AUTOMATE THIS TO SELECT BY LAST 3 DIGITS

lineseg = []
clic = []
#MAKE LIST OF LINE SEGMENTS BY ITERATING ROWS OVER ONLY COLUMN B
for row in sheet.iter_rows('B{}:B{}'.format(sheet.min_row,sheet.max_row)):
    for cell in row:
        lineseg.append(cell.value)
#MAKE A LIST OF CLIC NUMBERS BY ITERATING ROWS OVER ONLY COLUMN C
for row in sheet.iter_rows('C{}:C{}'.format(sheet.min_row,sheet.max_row)):
    for cell in row:
        clic.append(cell.value)

#print(lineseg)
#print(clic)

#zip THE TWO LISTS TOGETHER into list of tuples (you can manipulate to have a dash between ls and clic)
comb = list(zip(lineseg, clic))
#print(comb)

combed = []
xcltext = []

for i in comb:
    if i != (None, None):
        combed.append(str(i[0]).zfill(4) + '-' + str(i[1]))
del(combed[0:2])
for i in combed:
    xcltext.append(i.rstrip())
#print(xcltext)

#SEPERATE INTO DIFFERENT FINAL LISTS

#unfounddgntxt
for i in dgntext3:
    if i not in xcltext:
        unfounddgntxt.append(i)
    else:
        dgnfound.append(i)

for i in unfounddgntxt:
    if len(i) <= 11:
        unfounddgntxt2.append(i)

#unfoundxcltxt
for i in xcltext:
    if i not in dgntext3:
        unfoundxcltxt.append(i)
    else:
        xclfound.append(i)

#Remove duplicates in dgnfound and xclfound lists (by changing to set and back to list):
dgnfound = list(set(dgnfound))
xclfound = list(set(xclfound))

#Delete anything in unfoundxcltxt containing "main" or "fsac":
for i in unfounddgntxt2:
    if "MAIN" in i.upper():
        unfounddgntxt2.remove(i)
for i in unfounddgntxt2:
    if "FSAC" in i.upper():
        unfounddgntxt2.remove(i)

#SORT FINAL LISTS FOR COMPARISON
dgnfound.sort()
xclfound.sort()
unfounddgntxt2.sort()
unfoundxcltxt.sort()
#PRINT FINAL LISTS FOR COMPARISON
print()
print('Found in dgn:')
print(dgnfound)
print()
print('Found in excel:')
print(xclfound)
print()
print('NOT found in Excel that is in dgn:')
print(unfounddgntxt2)
print()
print('NOT found in dgn that is in excel:')
print(unfoundxcltxt)


#WRITE LISTS TO SEPARATE COLUMNS IN CSV FILE
d = [dgnfound, xclfound, unfounddgntxt2, unfoundxcltxt]
export_data = zip_longest(*d, fillvalue = '')
with open(nameFile + '.csv', 'w', encoding="ISO-8859-1", newline='') as myfile:
      wr = csv.writer(myfile)
      wr.writerow(("DGN", "XCL", "Not in Excel", "Not in DGN"))
      wr.writerows(export_data)
myfile.close()

leave = input("To exit, enter E")
if leave.upper() == 'E':
    exit()